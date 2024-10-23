from openpyxl import load_workbook
from jinja2 import Environment, FileSystemLoader, select_autoescape
import datetime
import argparse
import config
import gen_axi_fabric
import gen_ahb_fabric
import gen_axi2apb
import gen_axi2axi
import gen_axi2axi_lite
import gen_axi2bram
import gen_ahb2apb
import gen_ahb2sram
import csv
import os
import sys
import fnmatch

sys.path.insert(1, '../addr_dec_gen')
import addr_dec_gen

#################################################################
# Description         : Print list                              #
#################################################################
def print_list(lst):
    for x in lst:
        print(x)

#################################################################
# Description         : Generation of slave csv file based on   #
#                       parameters from Excel                   #
# Parameters          :                                         #
#   -dir_path         : The path to the addr_dec                #
#   -slave_n          : Number of slaves                        #
#   -master_n         : Number of masters                       #
#   -first_pos        : Number of the first line - addr_map -   #
#                       (default row 14)                        #
#   -slv_addr_offset  : Offset of the addr_map - slave side     #
#                       (default cell B (num 2))                #
#   -start_addr_offset: Offset of the addr_map - start addr     #
#                       (default cell C (num 3))                #
#   -end_addr_offset  : Offset of the addr_map - end addr       #
#                       (default cell D (num 4))                #
#################################################################
def create_slv_csv(dir_path, slave_n, master_n, first_pos,
                   slv_addr_offset, start_addr_offset,
                   end_addr_offset):
    tmp = []

    if slave_n != 0:
        slv_path = os.path.join(dir_path, "slave.csv")
        f = open(slv_path, 'w', newline='')
        writer = csv.writer(f)

        for x in range (0, master_n):
            tmp.append("range"+str(x))
        writer.writerow(tmp)

        for x in range (0, slave_n):
            if x == sheet.cell(first_pos+x,slv_addr_offset).value:
                writer.writerow([str(sheet.cell(first_pos+x,start_addr_offset).value) + "-" +
                                 str(sheet.cell(first_pos+x,end_addr_offset).value)])
            else:
                print("Error: range for slave", x, "does not exist.")
                exit()
        f.close()
    else:
        print("Error: must be at least 1 slave")
        exit()

#################################################################
# Description       : Generation of master csv file based on    #
#                     parameters from Excel                     #
# Parameters        :                                           #
#   -dir_path       : The path to the addr_dec                  #
#   -slave_n        : Number of slaves                          #
#   -master_n       : Number of masters                         #
#   -first_pos      : Number of the first line - visibility -   #
#                     master side (default row 22)              #
#   -mst_vis_offset : Offset of the visibility - master side    #
#                     (default cell B (num 2))                  #
#   -slv_vis_offset : Offset of visibility - slave side         #
#                     (default cell C (num 3))                  #
#################################################################
def create_mst_csv(dir_path, slave_n, master_n, first_pos,
                   mst_vis_offset, slv_vis_offset):
    tmp = []
    tmp_value = []

    mst_path = os.path.join(dir_path, "master.csv")
    f = open(mst_path, 'w', newline='')
    writer = csv.writer(f)

    if master_n != 0:
        for x in range (0, slave_n):
            tmp.append("slave"+str(x))
        writer.writerow(tmp)
        tmp.clear()

        for x in range (0, master_n):
            if x == sheet.cell(first_pos+x,mst_vis_offset).value:
                for j in range (0, slave_n):
                    tmp.append(str(sheet.cell(first_pos+x,slv_vis_offset+j).value))
                writer.writerow(tmp)
                tmp.clear()
            else:
                print("Error: visibility for master", x, "does not exist.")
                exit()
    else:
        for x in range (0, slave_n):
            tmp.append("slave"+str(x))
            tmp_value.append(str(1))

        writer.writerow(tmp)
        writer.writerow(tmp_value)
    f.close()

#################################################################
# Description      : Generating master and slave csv files and  #
#                    generating address decoders based on them  #
# Parameters          :                                         #
#   -dir_path         : The path to the TOP                     #
#   -addr_dec_name    : The name of the address decoder         #
#                       (axi_addr_dec_0, apb_addr_dec_0, ....)  #
#   -slave_n          : Number of slaves                        #
#   -master_n         : Number of masters                       #
#   -addr_first_pos   : Number of the first line - addr_map -   #
#                       (default row 14)                        #
#   -slv_addr_offset  : NOffset of the addr_map - slave side    #
#                       (default cell B (num 2))                #
#   -start_addr_offset: Offset of the addr_map - start addr     #
#                       (default cell C (num 3))                #
#   -end_addr_offset  : Offset of the addr_map - end addr       #
#                       (default cell D (num 4))                #
#   -vis_first_pos    : Number of the first line - visibility - #
#                       master side (default row 22)            #
#   -mst_vis_offset   : Offset of the visibility - master side  #
#                       (default cell B (num 2))                #
#   -slv_vis_offset   : Offset of visibility - slave side       #
#                       (default cell C (num 3))                #
#################################################################
def gen_addr_dec(dir_path, addr_dec_name, slave_n, master_n, addr_first_pos,
                 slv_addr_offset, start_addr_offset, end_addr_offset,
                 vis_first_pos, mst_vis_offset, slv_vis_offset):
    addr_dec_path = os.path.join(dir_path, addr_dec_name)
    if os.path.exists(addr_dec_path) == False:
        os.mkdir(addr_dec_path)
    create_slv_csv(addr_dec_path, slave_n, master_n, addr_first_pos,
                   slv_addr_offset, start_addr_offset, end_addr_offset)
    create_mst_csv(addr_dec_path, slave_n, master_n, vis_first_pos,
                   mst_vis_offset, slv_vis_offset)
    addr_dec_gen.generate_decoder(addr_dec_path, 1, 0, 0)

#################################################################
# Description       : Generate the axi_fabric module as well as #
#                     the corresponding address decoder based on#
#                     the parameters from Excel                 #
# Parameters        :                                           #
#   -sheet          : Excel sheet for current NOC               #
#                     (axi_fabric_0, axi_fabric_1....)          #
#   -dir_path       : The path to the TOP                       #
#   -addr_dec_name  : The name of the address decoder           #
#                     (apb_addr_dec_0, apb_addr_dec_1, ....)    #
#   -axi_fabric_name: The name of the axi_fabric module         #
#                     (axi_fabric_0, axi_fabric_1, ....)        #
#   -slvs_con       : List of slave connections                 #
#   -msts_con       : List of master connections                #
#   -options_def    : Default options for the axi_fabric module #
#   -param_def      : Default parameters for the axi_fabric     #
#   -first_pos      : Number of the first line with parameters  #
#                     (default row 2)                           #
#   -param_num      : Number of parameters                      #
#   -param_offset   : Parameters offset                         #
#   -value_offset   : Value offset                              #
#   -addr_first_pos   : Number of the first line - addr_map -   #
#                       (default row 14)                        #
#   -slv_addr_offset  : NOffset of the addr_map - slave side    #
#                       (default cell B (num 2))                #
#   -start_addr_offset: Offset of the addr_map - start addr     #
#                       (default cell C (num 3))                #
#   -end_addr_offset  : Offset of the addr_map - end addr       #
#                       (default cell D (num 4))                #
#   -vis_first_pos    : Number of the first line - visibility - #
#                       master side (default row 22)            #
#   -mst_vis_offset   : Offset of the visibility - master side  #
#                       (default cell B (num 2))                #
#   -slv_vis_offset   : Offset of visibility - slave side       #
#                       (default cell C (num 3))                #
#################################################################
def gen_axi_fabric_by_sheet(sheet, dir_path, addr_dec_name, axi_fabric_name,
                            slvs_con, msts_con, options_def, param_def, first_pos,
                            param_num, param_offset, value_offset, addr_first_pos,
                            slv_addr_offset, start_addr_offset, end_addr_offset,
                            vis_first_pos, mst_vis_offset, slv_vis_offset):

    options = []
    param_value = []

    # Get options from xlsx
    for x in sheet['A']:
        if x.value != None:
            options.append(x.value)

    # Check options
    for x in range (0, len(options_def)):
        if options[x] != options_def[x]:
            print("Error: all options for axi_fabric must be present")
            print_list(options_def)
            exit()

    # Check parameters
    for x in range (first_pos,first_pos+param_num):
        if sheet.cell(x,param_offset).value == param_def[x-param_offset]:
            if sheet.cell(x,value_offset).value != None:
                param_value.append(sheet.cell(x,value_offset).value)
            else:
                print("Error: The value of all parameters is mandatory")
                exit()
        else:
            print("Error: all parameters for axi_fabric must be present")
            print("Use the following order:")
            print_list(param_def)
            exit()

    global SLAVE_N, MASTER_N, ADDR_WIDTH, DATA_WIDTH, S_ID_WIDTH, M_ID_WIDTH, STRB_WIDTH
    global AW_USER_WIDTH, W_USER_WIDTH, B_USER_WIDTH, AR_USER_WIDTH, R_USER_WIDTH, MAX_TRANS
    global S_AR_CHANNEL_REG, M_AR_CHANNEL_REG, S_R_CHANNEL_REG, M_R_CHANNEL_REG, S_AW_CHANNEL_REG, M_AW_CHANNEL_REG
    global S_W_CHANNEL_REG, M_W_CHANNEL_REG, S_WR_CHANNEL_REG, M_WR_CHANNEL_REG

    ADDR_WIDTH          = param_value[0]
    DATA_WIDTH          = param_value[1]
    STRB_WIDTH          = param_value[2]
    AW_USER_WIDTH       = param_value[3]
    W_USER_WIDTH        = param_value[4]
    B_USER_WIDTH        = param_value[5]
    AR_USER_WIDTH       = param_value[6]
    R_USER_WIDTH        = param_value[7]
    S_ID_WIDTH          = param_value[8]
    M_ID_WIDTH          = param_value[9]
    ARB_Type            = param_value[10]
    MASTER_N            = param_value[11]
    SLAVE_N             = param_value[12]
    MAX_TRANS           = param_value[13]
    S_AR_CHANNEL_REG    = param_value[14]
    M_AR_CHANNEL_REG    = param_value[15]
    S_R_CHANNEL_REG     = param_value[16]
    M_R_CHANNEL_REG     = param_value[17]
    S_AW_CHANNEL_REG    = param_value[18]
    M_AW_CHANNEL_REG    = param_value[19]
    S_W_CHANNEL_REG     = param_value[20]
    M_W_CHANNEL_REG     = param_value[21]
    S_WR_CHANNEL_REG    = param_value[22]
    M_WR_CHANNEL_REG    = param_value[23]


    # Address decoder generation
    gen_addr_dec(dir_path, addr_dec_name, MASTER_N, SLAVE_N, addr_first_pos,
                 slv_addr_offset, start_addr_offset, end_addr_offset, vis_first_pos,
                 mst_vis_offset, slv_vis_offset)
    # Generating axi_fabric
    gen_axi_fabric.generate_axi_fabric(dir_path, axi_fabric_name, "axi_fabric_template.txt",
                                       SLAVE_N, MASTER_N, DATA_WIDTH, ADDR_WIDTH, STRB_WIDTH, AW_USER_WIDTH,
                                       W_USER_WIDTH, B_USER_WIDTH, AR_USER_WIDTH, R_USER_WIDTH,
                                       MAX_TRANS, S_ID_WIDTH, M_ID_WIDTH, S_AR_CHANNEL_REG, M_AR_CHANNEL_REG,
                                       S_R_CHANNEL_REG, M_R_CHANNEL_REG, S_AW_CHANNEL_REG,
                                       M_AW_CHANNEL_REG, S_W_CHANNEL_REG, M_W_CHANNEL_REG,
                                       S_WR_CHANNEL_REG, M_WR_CHANNEL_REG, slvs_con, msts_con)

#################################################################                       
# Description       : Generate the abh_fabric module as well as #
#                     the corresponding address decoder based on#
#                     the parameters from Excel                 #
# Parameters        :                                           #
#   -sheet          : Excel sheet for current NOC               #
#                     (ahb_fabric_0)                            #
#   -dir_path       : The path to the TOP                       #
#   -addr_dec_name  : The name of the address decoder           #
#                     (apb_addr_dec_0, apb_addr_dec_1, ....)    #
#   -axi_fabric_name: The name of the axi_fabric module         #
#                     (axi_fabric_0, axi_fabric_1, ....)        #
#   -slvs_con       : List of slave connections                 #
#   -msts_con       : List of master connections                #
#   -options_def    : Default options for the axi_fabric module #
#   -param_def      : Default parameters for the axi_fabric     #
#   -first_pos      : Number of the first line with parameters    #
#                     (default row 2)                           #
#   -param_num      : Number of parameters                      #
#   -param_offset   : Parameters offset                           #
#   -value_offset   : Value offset                                #
#   -addr_first_pos   : Number of the first line - addr_map -     #
#                       (default row 14)                        #
#   -slv_addr_offset  : NOffset of the addr_map - slave side      #
#                       (default cell B (num 2))                #
#   -start_addr_offset: Offset of the addr_map - start addr       #
#                       (default cell C (num 3))                #
#   -end_addr_offset  : Offset of the addr_map - end addr         #
#                       (default cell D (num 4))                #
#   -vis_first_pos    : Number of the first line - visibility -   #
#                       master side (default row 22)            #
#   -mst_vis_offset   : Offset of the visibility - master side    #
#                       (default cell B (num 2))                #
#   -slv_vis_offset   : Offset of visibility - slave side         #
#                       (default cell C (num 3))                #
#################################################################
def gen_ahb_fabric_by_sheet(sheet, dir_path, addr_dec_name, ahb_fabric_name,
                            slvs_con, msts_con, options_def, param_def, first_pos,
                            param_num, param_offset, value_offset, addr_first_pos,
                            slv_addr_offset, start_addr_offset, end_addr_offset,
                            vis_first_pos, mst_vis_offset, slv_vis_offset, slaves_range_min, slaves_range_max):

    options = []
    param_value = []

    # Get options from xlsx
    for x in sheet['A']:
        if x.value is not None:
            options.append(x.value)

    # Check options
    for x in range(len(options_def)):
        if options[x] != options_def[x]:
            print("Error: all options for ahb_fabric must be present")
            print_list(options_def)
            exit()

    # Check parameters
    for x in range(first_pos, first_pos + param_num):
        if sheet.cell(x, param_offset).value == param_def[x - param_offset]:
            if sheet.cell(x, value_offset).value is not None:
                param_value.append(sheet.cell(x, value_offset).value)
            else:
                print("Error: The value of all parameters is mandatory")
                exit()
        else:
            print("Error: all parameters for ahb_fabric must be present")
            print("Use the following order:")
            print_list(param_def)
            exit()

    global AHB_AW, APB_AW, AHB_DW, NUM_AHB_MASTERS, NUM_AHB_SLAVES, NUM_APB_SLAVES, NUM_BEATS_PER_M, NUM_BEATS_PER_M_MT

    # Extract parameters from the Excel sheet
    AHB_AW             = param_value[0]
    APB_AW             = param_value[1]
    AHB_DW             = param_value[2]
    NUM_AHB_MASTERS    = param_value[3]
    NUM_AHB_SLAVES     = param_value[4]
    NUM_APB_SLAVES     = param_value[5]
    NUM_BEATS_PER_M    = param_value[6]
    NUM_BEATS_PER_M_MT = param_value[7]    

    # Generating ahb_fabric
    gen_ahb_fabric.generate_ahb_fabric(dir_path, ahb_fabric_name, "ahb_fabric_template.txt",
                                       NUM_AHB_SLAVES, NUM_AHB_MASTERS, AHB_AW, AHB_AW, AHB_DW, NUM_APB_SLAVES,
                                       NUM_BEATS_PER_M, NUM_BEATS_PER_M_MT, slvs_con, msts_con, slaves_range_min, slaves_range_max)

###################################################################
# Description: Finds slave address ranges in the                  #
#              excel sheet and returns lists with the hard values.#                        
# Parameters:                                                     #
###################################################################
def read_excel_values(sheet, first_pos_of_ahb_param, ahb_param_number, ahb_param_offset, addr_first_pos, slv_addr_offset, start_addr_offset, end_addr_offset):
    slaves_range_min = []
    slaves_range_max = []

    for x in range(first_pos_of_ahb_param, ahb_param_number):
        cell_value = sheet.cell(x, ahb_param_offset).value
        if cell_value == "NUM_AHB_SLAVES":
            num_ahb_slaves = sheet.cell(x, ahb_param_offset + 1).value
            
    row = addr_first_pos  
    for i in range(num_ahb_slaves):
        row = addr_first_pos + i 
        slv_no = sheet.cell(row, slv_addr_offset).value  
        if slv_no is None or slv_no >= num_ahb_slaves:
            print(f"Number of AHB slaves should be {num_ahb_slaves}")
        else:
            start_addr = sheet.cell(row, start_addr_offset).value  
            end_addr = sheet.cell(row, end_addr_offset).value  
  
        # Append the start and end addresses to the respective lists
        slaves_range_min.append(start_addr)
        slaves_range_max.append(end_addr)

    return slaves_range_min, slaves_range_max
    

#################################################################
# Description       : Generate the axi2apb module as well as the#
#                     corresponding address decoder based on the#
#                     parameters from Excel                     #
# Parameters        :                                           #
#   -sheet          : Excel sheet for current NOC               #
#                     (axi2apb_0, axi2apb_1....)                #
#   -dir_path       : The path to the TOP                       #
#   -addr_dec_name  : The name of the address decoder           #
#                     (apb_addr_dec_0, apb_addr_dec_1, ....)    #
#   -axi2apb_name   : The name of the axi2apb module            #
#                     (axi2apb_0, axi2apb_1, ....)              #
#   -slvs_con       : List of slave connections                 #
#   -msts_con       : List of master connections                #
#   -options_def    : Default options for the axi2apb module    #
#   -param_def      : Default parameters for the axi2apb module #
#   -first_pos      : Number of the first line with parameters  #
#                     (default row 2)                           #
#   -param_num      : Number of parameters                      #
#   -param_offset   : Parameters offset                         #
#   -value_offset   : Value offset                              #
#   -addr_first_pos   : Number of the first line - addr_map -   #
#                       (default row 14)                        #
#   -slv_addr_offset  : NOffset of the addr_map - slave side    #
#                       (default cell B (num 2))                #
#   -start_addr_offset: Offset of the addr_map - start addr     #
#                       (default cell C (num 3))                #
#   -end_addr_offset  : Offset of the addr_map - end addr       #
#                       (default cell D (num 4))                #
#   -vis_first_pos    : Number of the first line - visibility - #
#                       master side (default row 22)            #
#   -mst_vis_offset   : Offset of the visibility - master side  #
#                       (default cell B (num 2))                #
#   -slv_vis_offset   : Offset of visibility - slave side       #
#                       (default cell C (num 3))                #
#################################################################
def gen_axi2apb_by_sheet(sheet, dir_path, addr_dec_name, axi2apb_name, column,
                         slvs_con, msts_con, options_def, param_def, first_pos,
                         param_num, param_offset, value_offset, addr_first_pos,
                         slv_addr_offset, start_addr_offset, end_addr_offset,
                         vis_first_pos, mst_vis_offset, slv_vis_offset):

    options = []
    param_value = []

    # Get options from xlsx
    for x in sheet[column]:
        if x.value != None:
            options.append(x.value)

    # Check options
    for x in range (0, len(options_def)):
        if options[x] != options_def[x]:
            print("Error: all options for axi2apb must be present")
            print_list(options_def)
            exit()

    # Check parameters
    for x in range (first_pos,first_pos+param_num):
        if sheet.cell(x,param_offset).value == param_def[x-param_offset]:
            if sheet.cell(x,value_offset).value != None:
                param_value.append(sheet.cell(x,value_offset).value)
            else:
                print("Error: The value of all parameters is mandatory")
                exit()
        else:
            print("Error: all parameters for axi2apb must be present")
            print("Use the following order:")
            print_list(param_def)
            exit()

    global AXI_AW,AXI_DW,AXI_IDW,AXI_STRB,AXI_USER,APB_AW,APB_DW,APB_STRB

    AXI_AW      = param_value[0]
    AXI_DW      = param_value[1]
    AXI_IDW     = param_value[2]
    AXI_STRB    = param_value[3]
    AXI_USER    = param_value[4]
    APB_AW      = param_value[5]
    APB_DW      = param_value[6]
    APB_STRB    = param_value[7]
    SLAVE_N     = param_value[8]

    # Address decoder generation
    gen_addr_dec(dir_path, addr_dec_name, SLAVE_N, 0, addr_first_pos,
                 slv_addr_offset, start_addr_offset, end_addr_offset,
                 vis_first_pos, mst_vis_offset, slv_vis_offset)
    # Generating axi2apb
    gen_axi2apb.generate_axi2apb(dir_path, axi2apb_name, "axi2apb_template.txt",
                                 SLAVE_N, AXI_IDW, AXI_DW, AXI_AW, AXI_STRB, AXI_USER, APB_AW,
                                 APB_DW, APB_STRB, slvs_con, msts_con)

#################################################################
# Description       : Generate the axi2axi module based on the  #
#                     parameters from Excel                     #
# Parameters        :                                           #
#   -sheet          : Excel sheet for current NOC               #
#                     (axi2axi_0, axi2axi_1....)                #
#   -dir_path       : The path to the TOP                       #
#################################################################
def gen_axi2axi_by_sheet(sheet, dir_path, axi2axi_name, column,
                         slvs_con, msts_con, options_def, param_def, first_pos,
                         param_num, param_offset, value_offset):

    options = []
    param_value = []

    # Get options from xlsx
    for x in sheet[column]:
        if x.value != None:
            options.append(x.value)

    # Check options
    for x in range (0, len(options_def)):
        if options[x] != options_def[x]:
            print("Error: all options for axi2axi must be present")
            print_list(options_def)
            exit()

    # Check parameters
    for x in range (first_pos,first_pos+param_num):
        if sheet.cell(x,param_offset).value == param_def[x-param_offset]:
            if sheet.cell(x,value_offset).value != None:
                param_value.append(sheet.cell(x,value_offset).value)
            else:
                print("Error: The value of all parameters is mandatory")
                exit()
        else:
            print("Error: all parameters for axi2axi must be present")
            print("Use the following order:")
            print_list(param_def)
            exit()

    global AXI2AXI_ADDR_WIDTH, AXI2AXI_S_DATA_WIDTH, AXI2AXI_M_DATA_WIDTH, AXI2AXI_S_ID_WIDTH, AXI2AXI_NUM_OF_OUTSTAND
    global AXI2AXI_M_ID_WIDTH, AXI2AXI_S_AWUSER_WIDTH, AXI2AXI_M_AWUSER_WIDTH, AXI2AXI_S_WUSER_WIDTH, AXI2AXI_M_WUSER_WIDTH
    global AXI2AXI_S_BUSER_WIDTH, AXI2AXI_M_BUSER_WIDTH, AXI2AXI_S_ARUSER_WIDTH, AXI2AXI_M_ARUSER_WIDTH, AXI2AXI_S_RUSER_WIDTH
    global AXI2AXI_M_RUSER_WIDTH, AXI2AXI_CDC, AXI2AXI_ALIGNED_UNALIGNED

    AXI2AXI_ADDR_WIDTH         = param_value[0]
    AXI2AXI_S_DATA_WIDTH       = param_value[1]
    AXI2AXI_M_DATA_WIDTH       = param_value[2]
    AXI2AXI_S_ID_WIDTH         = param_value[3]
    AXI2AXI_NUM_OF_OUTSTAND    = param_value[4]
    AXI2AXI_M_ID_WIDTH         = param_value[5]
    AXI2AXI_S_AWUSER_WIDTH     = param_value[6]
    AXI2AXI_M_AWUSER_WIDTH     = param_value[7]
    AXI2AXI_S_WUSER_WIDTH      = param_value[8]
    AXI2AXI_M_WUSER_WIDTH      = param_value[9]
    AXI2AXI_S_BUSER_WIDTH      = param_value[10]
    AXI2AXI_M_BUSER_WIDTH      = param_value[11]
    AXI2AXI_S_ARUSER_WIDTH     = param_value[12]
    AXI2AXI_M_ARUSER_WIDTH     = param_value[13]
    AXI2AXI_S_RUSER_WIDTH      = param_value[14]
    AXI2AXI_M_RUSER_WIDTH      = param_value[15]
    AXI2AXI_CDC                = param_value[16]
    AXI2AXI_ALIGNED_UNALIGNED  = param_value[17]

    # Generating axi2axi
    gen_axi2axi.generate_axi2axi(dir_path, axi2axi_name, "axi2axi_template.txt",
                                 AXI2AXI_ADDR_WIDTH, AXI2AXI_S_DATA_WIDTH, AXI2AXI_M_DATA_WIDTH, AXI2AXI_S_ID_WIDTH, AXI2AXI_NUM_OF_OUTSTAND,
                                 AXI2AXI_M_ID_WIDTH, AXI2AXI_S_AWUSER_WIDTH, AXI2AXI_M_AWUSER_WIDTH, AXI2AXI_S_WUSER_WIDTH,
                                 AXI2AXI_M_WUSER_WIDTH, AXI2AXI_S_BUSER_WIDTH, AXI2AXI_M_BUSER_WIDTH, AXI2AXI_S_ARUSER_WIDTH,
                                 AXI2AXI_M_ARUSER_WIDTH, AXI2AXI_S_RUSER_WIDTH, AXI2AXI_M_RUSER_WIDTH, AXI2AXI_CDC,
                                 AXI2AXI_ALIGNED_UNALIGNED, msts_con, slvs_con)

#################################################################
# Description       : Generate the axi2axi_lite module based on #
#                     the parameters from Excel                 #
# Parameters        :                                           #
#   -sheet          : Excel sheet for current NOC               #
#                     (axi2axi_lite_0, axi2axi_lite_1....)      #
#   -dir_path       : The path to the TOP                       #
#################################################################
def gen_axi2axi_lite_by_sheet(sheet, dir_path, axi2axi_lite_name, column,
                         slvs_con, msts_con, options_def, param_def, first_pos,
                         param_num, param_offset, value_offset):

    options = []
    param_value = []

    # Get options from xlsx
    for x in sheet[column]:
        if x.value != None:
            options.append(x.value)

    # Check options
    for x in range (0, len(options_def)):
        if options[x] != options_def[x]:
            print("Error: all options for axi2axi_lite must be present")
            print_list(options_def)
            exit()

    # Check parameters
    for x in range (first_pos,first_pos+param_num):
        if sheet.cell(x,param_offset).value == param_def[x-param_offset]:
            if sheet.cell(x,value_offset).value != None:
                param_value.append(sheet.cell(x,value_offset).value)
            else:
                print("Error: The value of all parameters is mandatory")
                exit()
        else:
            print("Error: all parameters for axi2axi_lite must be present")
            print("Use the following order:")
            print_list(param_def)
            exit()

    global AXI2AXI_lite_ADDR_WIDTH, AXI2AXI_lite_S_DATA_WIDTH, AXI2AXI_lite_M_DATA_WIDTH, AXI2AXI_lite_S_ID_WIDTH, AXI2AXI_lite_NUM_OF_OUTSTAND
    global AXI2AXI_lite_S_AWUSER_WIDTH, AXI2AXI_lite_S_WUSER_WIDTH, AXI2AXI_lite_S_BUSER_WIDTH, AXI2AXI_lite_S_ARUSER_WIDTH
    global AXI2AXI_lite_S_RUSER_WIDTH, AXI2AXI_lite_CDC, AXI2AXI_lite_ALIGNED_UNALIGNED

    AXI2AXI_lite_ADDR_WIDTH         = param_value[0]
    AXI2AXI_lite_S_DATA_WIDTH       = param_value[1]
    AXI2AXI_lite_M_DATA_WIDTH       = param_value[2]
    AXI2AXI_lite_S_ID_WIDTH         = param_value[3]
    AXI2AXI_lite_NUM_OF_OUTSTAND    = param_value[4]
    AXI2AXI_lite_S_AWUSER_WIDTH     = param_value[5]
    AXI2AXI_lite_S_WUSER_WIDTH      = param_value[6]
    AXI2AXI_lite_S_BUSER_WIDTH      = param_value[7]
    AXI2AXI_lite_S_ARUSER_WIDTH     = param_value[8]
    AXI2AXI_lite_S_RUSER_WIDTH      = param_value[9]
    AXI2AXI_lite_CDC                = param_value[10]
    AXI2AXI_lite_ALIGNED_UNALIGNED  = param_value[11]

    # Generating AXI2AXI_lite
    gen_axi2axi_lite.generate_axi2axi_lite(dir_path, axi2axi_lite_name, "axi2axi_lite_template.txt",
                                           AXI2AXI_lite_ADDR_WIDTH, AXI2AXI_lite_S_DATA_WIDTH, AXI2AXI_lite_M_DATA_WIDTH, AXI2AXI_lite_S_ID_WIDTH, AXI2AXI_lite_NUM_OF_OUTSTAND,
                                           AXI2AXI_lite_S_AWUSER_WIDTH, AXI2AXI_lite_S_WUSER_WIDTH, AXI2AXI_lite_S_BUSER_WIDTH, AXI2AXI_lite_S_ARUSER_WIDTH,
                                           AXI2AXI_lite_S_RUSER_WIDTH, AXI2AXI_lite_CDC, AXI2AXI_lite_ALIGNED_UNALIGNED, msts_con, slvs_con)

#################################################################
# Description       : Generate the axi2bram module based on the #
#                     parameters from Excel                     #
# Parameters        :                                           #
#   -sheet          : Excel sheet for current NOC               #
#                     (axi2bram_0, axi2bram_1....)              #
#   -dir_path       : The path to the TOP                       #
#################################################################
def gen_axi2bram_by_sheet(sheet, dir_path, axi2bram_name, column,
                         slvs_con, msts_con, options_def, param_def, first_pos,
                         param_num, param_offset, value_offset):

    options = []
    param_value = []

    # Get options from xlsx
    for x in sheet[column]:
        if x.value != None:
            options.append(x.value)

    # Check options
    for x in range (0, len(options_def)):
        if options[x] != options_def[x]:
            print("Error: all options for axi2bram must be present")
            print_list(options_def)
            exit()

    # Check parameters
    for x in range (first_pos,first_pos+param_num):
        if sheet.cell(x,param_offset).value == param_def[x-param_offset]:
            if sheet.cell(x,value_offset).value != None:
                param_value.append(sheet.cell(x,value_offset).value)
            else:
                print("Error: The value of all parameters is mandatory")
                exit()
        else:
            print("Error: all parameters for axi2bram must be present")
            print("Use the following order:")
            print_list(param_def)
            exit()

    global AXI2AXI_ADDR_WIDTH, AXI2BRAM_ADDR_WIDTH, AXI2BRAM_S_DATA_WIDTH, AXI2BRAM_MEM_DEPTH, AXI2BRAM_S_ID_WIDTH        
    global AXI2BRAM_NUM_OF_OUTSTAND, AXI2BRAM_S_AWUSER_WIDTH, AXI2BRAM_S_WUSER_WIDTH, AXI2BRAM_S_BUSER_WIDTH, AXI2BRAM_S_ARUSER_WIDTH    
    global AXI2BRAM_S_RUSER_WIDTH, AXI2BRAM_ALIGNED_UNALIGNED, AXI2BRAM_MEM_INTERFACE_TYPE, AXI2BRAM_MEM_OPERATION_TYPE, AXI2BRAM_REGCEA            
    global AXI2BRAM_REGCEB, AXI2BRAM_PORT_EN_A, AXI2BRAM_PORT_EN_B         

    AXI2BRAM_ADDR_WIDTH          = param_value[0]
    AXI2BRAM_S_DATA_WIDTH        = param_value[1]
    AXI2BRAM_MEM_DEPTH           = param_value[2]
    AXI2BRAM_S_ID_WIDTH          = param_value[3]
    AXI2BRAM_NUM_OF_OUTSTAND     = param_value[4]
    AXI2BRAM_S_AWUSER_WIDTH      = param_value[5]
    AXI2BRAM_S_WUSER_WIDTH       = param_value[6]
    AXI2BRAM_S_BUSER_WIDTH       = param_value[7]
    AXI2BRAM_S_ARUSER_WIDTH      = param_value[8]
    AXI2BRAM_S_RUSER_WIDTH       = param_value[9]
    AXI2BRAM_ALIGNED_UNALIGNED   = param_value[10]
    AXI2BRAM_MEM_INTERFACE_TYPE  = param_value[11]
    AXI2BRAM_MEM_OPERATION_TYPE  = param_value[12]
    AXI2BRAM_REGCEA              = param_value[13]
    AXI2BRAM_REGCEB              = param_value[14]
    AXI2BRAM_PORT_EN_A           = param_value[15]
    AXI2BRAM_PORT_EN_B           = param_value[16]

    # Generating axi2axi
    gen_axi2bram.generate_axi2bram(dir_path, axi2bram_name, "axi2bram_template.txt",
                                  AXI2BRAM_ADDR_WIDTH, AXI2BRAM_S_DATA_WIDTH, AXI2BRAM_MEM_DEPTH, AXI2BRAM_S_ID_WIDTH, AXI2BRAM_NUM_OF_OUTSTAND, 
                                  AXI2BRAM_S_AWUSER_WIDTH, AXI2BRAM_S_WUSER_WIDTH, AXI2BRAM_S_BUSER_WIDTH, AXI2BRAM_S_ARUSER_WIDTH, AXI2BRAM_S_RUSER_WIDTH, 
                                  AXI2BRAM_ALIGNED_UNALIGNED, AXI2BRAM_MEM_INTERFACE_TYPE, AXI2BRAM_MEM_OPERATION_TYPE, AXI2BRAM_REGCEA, AXI2BRAM_REGCEB, 
                                  AXI2BRAM_PORT_EN_A, AXI2BRAM_PORT_EN_B, msts_con, slvs_con)

#################################################################
# Description       : Generate the ahb2apb module based on the  #
#                     parameters from Excel                     #
# Parameters        :                                           #
#   -sheet          : Excel sheet for current NOC               #
#                     (ahb2apb_0, ahb2apb_1....)                #
#   -dir_path       : The path to the TOP                       #
#################################################################
def gen_ahb2apb_by_sheet(sheet, dir_path, ahb2apb_name, column,
                         slvs_con, msts_con, options_def, param_def, first_pos,
                         param_num, param_offset, value_offset):

    options = []
    param_value = []

    # Get options from xlsx
    for x in sheet[column]:
        if x.value != None:
            options.append(x.value)

    # Check options
    for x in range (0, len(options_def)):
        if options[x] != options_def[x]:
            print("Error: all options for axi2axi must be present")
            print_list(options_def)
            exit()

    # Check parameters
    for x in range (first_pos,first_pos+param_num):
        if sheet.cell(x,param_offset).value == param_def[x-param_offset]:
            if sheet.cell(x,value_offset).value != None:
                param_value.append(sheet.cell(x,value_offset).value)
            else:
                print("Error: The value of all parameters is mandatory")
                exit()
        else:
            print("Error: all parameters for ahb2apb must be present")
            print("Use the following order:")
            print_list(param_def)
            exit()

    global AHB_AW, APB_AW, AHB_DW, APB_DW, FIFO_SIZE, AHB_ENDIAN, APB_ENDIAN

    AHB_AW     = param_value[0]
    APB_AW     = param_value[1]
    AHB_DW     = param_value[2]
    APB_DW     = param_value[3]
    FIFO_SIZE  = param_value[4]
    AHB_ENDIAN = param_value[5]
    APB_ENDIAN = param_value[6]

    # Generating ahb2apb
    gen_ahb2apb.generate_ahb2apb(dir_path, ahb2apb_name, "ahb2apb_template.txt",
                                 AHB_AW, APB_AW, AHB_DW, APB_DW, FIFO_SIZE, AHB_ENDIAN, 
                                 APB_ENDIAN, slvs_con, msts_con)

#################################################################
# Description       : Generate the ahb2sram module based on the #
#                     parameters from Excel                     #
# Parameters        :                                           #
#   -sheet          : Excel sheet for current NOC               #
#                     (ahb2sram_0, ahb2sram_1....)              #
#   -dir_path       : The path to the TOP                       #
#################################################################
def gen_ahb2sram_by_sheet(sheet, dir_path, ahb2sram_name, column,
                         slvs_con, msts_con, options_def, param_def, first_pos,
                         param_num, param_offset, value_offset):

    options = []
    param_value = []

    # Get options from xlsx
    for x in sheet[column]:
        if x.value != None:
            options.append(x.value)

    # Check options
    for x in range (0, len(options_def)):
        if options[x] != options_def[x]:
            print("Error: all options for ahb2sram must be present")
            print_list(options_def)
            exit()

    # Check parameters
    for x in range (first_pos,first_pos+param_num):
        if sheet.cell(x,param_offset).value == param_def[x-param_offset]:
            if sheet.cell(x,value_offset).value != None:
                param_value.append(sheet.cell(x,value_offset).value)
            else:
                print("Error: The value of all parameters is mandatory")
                exit()
        else:
            print("Error: all parameters for ahb2sram must be present")
            print("Use the following order:")
            print_list(param_def)
            exit()

    global DATA_WIDTH, MEM_DEPTH         

    DATA_WIDTH = param_value[0]
    MEM_DEPTH  = param_value[1]

    # Generating axi2axi
    gen_ahb2sram.generate_ahb2sram(dir_path, ahb2sram_name, "ahb2sram_template.txt",
                                   DATA_WIDTH, MEM_DEPTH, msts_con, slvs_con)

#################################################################
# Description    : Adds an interface or wire to the TOP file     #
#                  depending on the connections and template.   #
# Parameters     :                                              #
#   -dir_path    : The path to the file                          #
#   -file_name    : The name of the file to write to              #
#   -mst_slv     : Master or slave connections                  #
#   -template    : A template of what the file should look like  #
#################################################################
def add_interface_to_top(dir_path, file_name, mst_slv, template):

    date = datetime.datetime.now()
    template = env.get_template(template)
    template_r = template.render(module_name=file_name, mst_slv=mst_slv, date=date)

    with open(dir_path+file_name+'.sv', 'a') as file_handler:
        file_handler.write(template_r)

#################################################################
# Description    :                                              #
#                                                               #
# Parameters     :                                              #
#   -dir_path    :                                              #
#   -file_name   :                                              #
#   -mst_slv     :                                              #
#   -template    :                                              #
#################################################################
def end_interface(dir_path, file_name):

    with open(dir_path+file_name+'.sv', 'a') as file_handler:
        file_handler.write(");")
    replace(dir_path+top_name+".sv", ",);"      , "\n);\n")

#################################################################
# Description    :                                              #
#                                                               #
# Parameters     :                                              #
#   -dir_path    :                                              #
#   -file_name   :                                              #
#   -mst_slv     :                                              #
#   -template    :                                              #
#################################################################
def add_clk_rst_to_top(dir_path, file_name, noc_name, template):

    template = env.get_template(template)
    template_r = template.render(noc_name=noc_name)

    with open(dir_path+file_name+'.sv', 'a') as file_handler:
        file_handler.write(template_r)

#################################################################
# Description    :                                              #
#                                                               #
# Parameters     :                                              #
#   -dir_path    :                                              #
#   -file_name   :                                              #
#   -mst_slv     :                                              #
#   -template    :                                              #
#################################################################
def add_clk_rst_to_interface(dir_path, top_name, noc_name):
    if "AXI_fabric" in noc_name:
        add_clk_rst_to_top(dir_path, top_name, noc_name+"_wrapper", "axi_clk_rst_template.txt")
    elif "AHB_fabric" in noc_name:
        add_clk_rst_to_top(dir_path, top_name, noc_name+"_wrapper", "ahb_clk_rst_template.txt")
    elif "axi2apb" in noc_name:
        add_clk_rst_to_top(dir_path, top_name, noc_name+"_wrapper", "apb2axi_clk_rst_template.txt")
    elif "axi2axi" in noc_name:
        add_clk_rst_to_top(dir_path, top_name, noc_name+"_wrapper", "axi2axi_clk_rst_template.txt")
    elif "axi2axi_lite" in noc_name:
        add_clk_rst_to_top(dir_path, top_name, noc_name+"_wrapper", "axi2axi_clk_rst_template.txt")
    elif "axi2bram" in noc_name:
        add_clk_rst_to_top(dir_path, top_name, noc_name+"_wrapper", "axi2bram_clk_rst_template.txt")
    elif "ahb2apb" in noc_name:
        add_clk_rst_to_top(dir_path, top_name, noc_name+"_wrapper", "ahb2apb_clk_rst_template.txt")  
    elif "ahb2sram" in noc_name:
        add_clk_rst_to_top(dir_path, top_name, noc_name+"_wrapper", "ahb2sram_clk_rst_template.txt")           
    else:
        print("Error: wrong NOC")
        exit()

#################################################################
# Description    : Finds all master connections for the current #
#                  NOC. In case the connection is               #
#                  input (first letter 'm') add the required     #
#                  interface to the TOP file.                    #
#                  Otherwise, it stores the connection as a wire#
# Parameters     :                                              #
#   -dir_path    : The path to the TOP                          #
#   -cnt         : The sequence  number of the current NOC      #
#   -sheet       : Sheet with connections                       #
#   -mst_offset  : Offset of the master relative to the           #
#                  NOC cell (default cell C (num 3))            #
#   -top_name    : TOP file name                                 #
#   -first_pos   : Number of the first line with connections      #
#                  (default row 2)                              #
#   -max_con_num : Maximum number of connections                #
#                  (default row 16)                             #
#                                                               #
# Returns        : List of master connections                   #
#################################################################
def get_mst_con(dir_path, cnt, sheet, mst_offset, top_name, first_pos, max_con_num, noc):
    master_connection = []
    is_wire = False
    for i in range (first_pos, first_pos+max_con_num):
        mst = sheet.cell(i,cnt*3+mst_offset).value
        if mst != None:
            master_connection.append(mst)
            if mst[0] == "m":
                if "axi" in mst:
                    if "axi2axi_lite" in noc:  # TODO update
                        add_interface_to_top(dir_path, top_name, mst, "axi2axi_lite_slave_template.txt")   
                    elif "axi2axi" in noc:
                        add_interface_to_top(dir_path, top_name, mst, "axi2axi_slave_template.txt")
                    elif "axi2bram" in noc:
                        add_interface_to_top(dir_path, top_name, mst, "axi2bram_slave_template.txt")
                    elif "ahb2sram" in noc:
                        add_interface_to_top(dir_path, "TOP_wire", mst, "ahb2sram_slave_template.txt") 
                    else:
                        add_interface_to_top(dir_path, top_name, mst, "axi_slave_template.txt")
                elif "apb" in mst:
                    add_interface_to_top(dir_path, top_name, mst, "apb_slave_template.txt")
                elif "ahb" in mst:
                    add_interface_to_top(dir_path, top_name, mst, "ahb_slave_template.txt")
            elif mst[0] != "s":
                if "AXI_fabric" in noc:
                    add_interface_to_top(dir_path, "TOP_wire", mst, "wire_axi_fabric_template.txt")
                elif "ahb2apb" in noc:
                    add_interface_to_top(dir_path, "TOP_wire", mst, "wire_ahb_fabric_template.txt")
                elif "axi2bram" in noc:
                    add_interface_to_top(dir_path, "TOP_wire", mst, "wire_axi2bram_template.txt")
                elif "ahb2sram" in noc:
                    add_interface_to_top(dir_path, "TOP_wire", mst, "wire_ahb2sram_template.txt")                   
                else:
                    add_interface_to_top(dir_path, "TOP_wire", mst, "wire_template.txt")
                is_wire = True
        else:
            break
    return master_connection, is_wire

#################################################################
# Description    : Finds all slave connections for the current  #
#                  NOC. In case the connection is               #
#                  output (first letter 's') add the required    #
#                  interface to the TOP file.                    #
# Parameters     :                                              #
#   -dir_path    : The path to the TOP                          #
#   -cnt         : The sequence  number of the current NOC      #
#   -sheet       : Sheet with connections                       #
#   -slv_offset  : Offset of the slave relative to the            #
#                  NOC cell (default cell D (num 4))            #
#   -top_name    : TOP file name                                 #
#   -first_pos   : Number of the first line with connections      #
#                  (default row 2)                              #
#   -max_con_num : Maximum number of connections                #
#                  (default row 16)                             #
#                                                               #
# Returns        : List of slave connections                    #
#################################################################
def get_slv_con(dir_path, cnt, sheet, slv_offset, top_name, first_pos, max_con_num, noc):
    slave_connection = []
    for i in range (first_pos, first_pos+max_con_num):
        slv = sheet.cell(i,cnt*3+slv_offset).value
        if slv != None:
            slave_connection.append(slv)
            if slv[0] == "s":
                if "axi" in slv:
                    if "axi2axi_lite" in noc:    
                        add_interface_to_top(dir_path, top_name, slv, "axi2axi_lite_master_template.txt")
                    elif "axi2axi" in noc:
                        add_interface_to_top(dir_path, top_name, slv, "axi2axi_master_template.txt")
                    else:
                        add_interface_to_top(dir_path, top_name, slv, "axi_master_template.txt")
                elif "apb" in slv:
                    add_interface_to_top(dir_path, top_name, slv, "apb_master_template.txt")
                elif "ahb" in slv:
                    if "ahb2apb" in noc:
                        add_interface_to_top(dir_path, top_name, slv, "ahb2apb_master_template.txt")
                    else:
                        add_interface_to_top(dir_path, top_name, slv, "ahb_master_template.txt")
        else:
            break
    return slave_connection

#################################################################
# Description    : Replaces the string in the file with the new #
#                  given string                                 #
# Parameters     :                                              #
#   -des_path    : The path to the file                         #
#   -old         : Old string                                   #
#   -new         : New string                                   #
#################################################################
def replace(des_path, old, new):
    f = open(des_path, "rt")
    data = f.read()
    data = data.replace(old, new)
    f.close()
    f = open(des_path, "wt")
    f.write(data)
    f.close()

#################################################################
# Description    : Copies the original AXI files and adapts     #
#                  them to the new address decoder              #
#                  (different values for different axi_fabrics) #
# Parameters     :                                              #
#   -org_path    : Path to original AXI files                   #
#   -new_path    : Path to new AXI files                        #
#   -cnt         : Sequence number of axi_fabric                #
#################################################################
def copy_and_update_axi_file(org_path, new_path, cnt):
    os.mkdir(new_path)
    # copy AXI files
    os.system("cp "+org_path+"axi_fabric_top.sv "+new_path)
    os.system("cp "+org_path+"addr_dec_mst_sel.sv "+new_path)
    # Rename the files to contain the sequence number NOC
    os.rename(new_path+"/axi_fabric_top.sv", new_path+"/axi_fabric_top_"+str(cnt)+".sv")
    os.rename(new_path+"/addr_dec_mst_sel.sv", new_path+"/addr_dec_mst_sel_"+str(cnt)+".sv")        
    # Update the files with the new address decoder name and NOC sequence number
    replace(new_path+"/axi_fabric_top_"+str(cnt)+".sv", "axi_fabric_top", "axi_fabric_top_"+str(cnt))
    replace(new_path+"/axi_fabric_top_"+str(cnt)+".sv", "addr_dec_mst_sel", "addr_dec_mst_sel_"+str(cnt))
    replace(new_path+"/addr_dec_mst_sel_"+str(cnt)+".sv", "addr_dec_mst_sel", "addr_dec_mst_sel_"+str(cnt))
    replace(new_path+"/addr_dec_mst_sel_"+str(cnt)+".sv", "addr_soc", "axi_addr_dec_"+str(cnt))

#################################################################
# Description    : Copies the original AHB files and adapts     #
#                  them to the new address decoder              #
#                  (different values for different axi_fabrics) #
# Parameters     :                                              #
#   -org_path    : Path to original AHB files                   #
#   -new_path    : Path to new AHB files                        #
#   -cnt         : Sequence number of ahb_fabric                #
#################################################################
def copy_and_update_ahb_file(org_path, new_path, cnt):
    os.mkdir(new_path)
    # Copy AHB files
    os.system("cp "+org_path+"ahb_arbiter_top.sv "+new_path)
    os.system("cp "+org_path+"address_decoder_mix.sv "+new_path)      # need to copy them also, add sequence number for each new noc
    # Rename the files with the sequence number
    os.rename(new_path+"/ahb_arbiter_top.sv", new_path+"/ahb_arbiter_top_"+str(cnt)+".sv")
    os.rename(new_path+"/address_decoder_mix.sv", new_path+"/address_decoder_mix_"+str(cnt)+".sv")
    # Update the files with the new names and sequence numbers
    replace(new_path+"/ahb_arbiter_top_"+str(cnt)+".sv", "ahb_arbiter_top", "ahb_arbiter_top_"+str(cnt))
    replace(new_path+"/ahb_arbiter_top_"+str(cnt)+".sv", "address_decoder_mix", "address_decoder_mix_"+str(cnt))
    replace(new_path+"/address_decoder_mix_"+str(cnt)+".sv", "addr_soc", "ahb_addr_dec_"+str(cnt))

#################################################################
# Description    : Copies the original AXI2APB files and adapts #
#                  them to the new address decoder              #
#                  (different values for different axi2apb)     #
# Parameters     :                                              #
#   -org_path    : Path to original APB files                   #
#   -new_path    : Path to new APB files                        #
#   -cnt         : Sequence number of axi2apb                   #
#################################################################
def copy_and_update_apb_file(org_path, new_path, cnt):
    os.mkdir(new_path)
    # copy APB files
    os.system("cp "+org_path+"axi2apb.sv "+new_path)
    # Rename the files to contain the sequence number NOC
    os.rename(new_path+"/axi2apb.sv", new_path+"/axi2apb_"+str(cnt)+".sv")
    # Update the files with the new address decoder name and NOC sequence number
    replace(new_path+"/axi2apb_"+str(cnt)+".sv", "axi2apb", "axi2apb_"+str(cnt))
    replace(new_path+"/axi2apb_"+str(cnt)+".sv", "addr_dec", "apb_addr_dec_"+str(cnt))

#################################################################
# Description    :                                              #
#                                                               #
#                                                               #
# Parameters     :                                              #
#################################################################
def replace_axi_fabric_param(axi_fabric_path, axi_cnt, sv_dir, top_name):
    replace(axi_fabric_path+"/AXI_fabric_"+str(axi_cnt)+"_wrapper.sv", "axi_fabric_top_tmp", "axi_fabric_top_"+str(axi_cnt))
    replace(sv_dir+top_name+".sv", "DATA_WIDTH", str(DATA_WIDTH))
    replace(sv_dir+top_name+".sv", "ADDR_WIDTH", str(ADDR_WIDTH))
    replace(sv_dir+top_name+".sv", "STRB_WIDTH", str(STRB_WIDTH))
    replace(sv_dir+top_name+".sv", "AW_USER_WIDTH", str(AW_USER_WIDTH))
    replace(sv_dir+top_name+".sv", "W_USER_WIDTH", str(W_USER_WIDTH))
    replace(sv_dir+top_name+".sv", "B_USER_WIDTH", str(B_USER_WIDTH))
    replace(sv_dir+top_name+".sv", "AR_USER_WIDTH", str(AR_USER_WIDTH))
    replace(sv_dir+top_name+".sv", "R_USER_WIDTH", str(R_USER_WIDTH))
    replace(sv_dir+top_name+".sv", "S_ID_WIDTH",   str(S_ID_WIDTH))
    replace(sv_dir+top_name+".sv", "M_ID_WIDTH",   str(M_ID_WIDTH))

#################################################################
# Description    :                                              #
#                                                               #
#                                                               #
# Parameters     :                                              #
#################################################################

def replace_ahb_fabric_param(ahb_fabric_path, ahb_cnt, sv_dir, top_name):
    replace(ahb_fabric_path+"/AHB_fabric_"+str(ahb_cnt)+"_wrapper.sv", "ahb_arbiter_top_tmp", "ahb_arbiter_top_"+str(ahb_cnt)) 
    replace(sv_dir+top_name+".sv", "AHB_AW", str(AHB_AW))
    replace(sv_dir+top_name+".sv", "APB_AW", str(APB_AW))
    replace(sv_dir+top_name+".sv", "AHB_DW", str(AHB_DW))
    replace(sv_dir+top_name+".sv", "NUM_AHB_MASTERS", str(NUM_AHB_MASTERS))
    replace(sv_dir+top_name+".sv", "NUM_AHB_SLAVES", str(NUM_AHB_SLAVES))   
    replace(sv_dir+top_name+".sv", "NUM_APB_SLAVES", str(NUM_APB_SLAVES))
    replace(sv_dir+top_name+".sv", "NUM_BEATS_PER_M", str(NUM_BEATS_PER_M))
    replace(sv_dir+top_name+".sv", "NUM_BEATS_PER_M_MT",   str(NUM_BEATS_PER_M_MT))

#################################################################
# Description    :                                              #
#                                                               #
#                                                               #
# Parameters     :                                              #
#################################################################
def replace_axi2apb_param(axi2apb_path, apb_cnt, sv_dir, top_name):
    replace(axi2apb_path+"/axi2apb_"+str(apb_cnt)+"_wrapper.sv", "axi2apb_tmp", "axi2apb_"+str(apb_cnt))
    replace(sv_dir+top_name+".sv", "ADDR_WIDTH", str(AXI_AW))
    replace(sv_dir+top_name+".sv", "DATA_WIDTH", str(AXI_DW))
    replace(sv_dir+top_name+".sv", "S_ID_WIDTH",   str(AXI_IDW))
    replace(sv_dir+top_name+".sv", "STRB_WIDTH", str(AXI_STRB))
    replace(sv_dir+top_name+".sv", "AW_USER_WIDTH", str(AXI_USER))
    replace(sv_dir+top_name+".sv", "W_USER_WIDTH", str(AXI_USER))
    replace(sv_dir+top_name+".sv", "B_USER_WIDTH", str(AXI_USER))
    replace(sv_dir+top_name+".sv", "AR_USER_WIDTH", str(AXI_USER))
    replace(sv_dir+top_name+".sv", "R_USER_WIDTH", str(AXI_USER))
    replace(sv_dir+top_name+".sv", "APB_addr_length", str(APB_AW))
    replace(sv_dir+top_name+".sv", "APB_data_length", str(APB_DW))
    replace(sv_dir+top_name+".sv", "APB_strb_length", str(APB_STRB))

#################################################################
# Description    :                                              #
#                                                               #
#                                                               #
# Parameters     :                                              #
#################################################################
def replace_axi2axi_param(axi2axi_path, axi2axi_cnt, sv_dir, top_name):
    replace(sv_dir+top_name+".sv", "ADDR_WIDTH"      , str(AXI2AXI_ADDR_WIDTH))
    replace(sv_dir+top_name+".sv", "S_DATA_WIDTH"    , str(AXI2AXI_S_DATA_WIDTH))
    replace(sv_dir+top_name+".sv", "S_ID_WIDTH"      , str(AXI2AXI_S_ID_WIDTH))
    replace(sv_dir+top_name+".sv", "S_AWUSER_WIDTH"  , str(AXI2AXI_S_AWUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "S_WUSER_WIDTH"   , str(AXI2AXI_S_WUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "S_BUSER_WIDTH"   , str(AXI2AXI_S_BUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "S_ARUSER_WIDTH"  , str(AXI2AXI_S_ARUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "S_RUSER_WIDTH"   , str(AXI2AXI_S_RUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "M_DATA_WIDTH"    , str(AXI2AXI_M_DATA_WIDTH))
    replace(sv_dir+top_name+".sv", "M_STRB_WIDTH"    , str(int(AXI2AXI_M_DATA_WIDTH/8)))
    replace(sv_dir+top_name+".sv", "M_ID_WIDTH"      , str(AXI2AXI_M_ID_WIDTH))
    replace(sv_dir+top_name+".sv", "M_AW_USER_WIDTH"  , str(AXI2AXI_M_AWUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "M_W_USER_WIDTH"   , str(AXI2AXI_M_WUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "M_B_USER_WIDTH"   , str(AXI2AXI_M_BUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "M_AR_USER_WIDTH"  , str(AXI2AXI_M_ARUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "M_R_USER_WIDTH"   , str(AXI2AXI_M_RUSER_WIDTH))
    
#################################################################
# Description    :                                              #
#                                                               #
#                                                               #
# Parameters     :                                              #
#################################################################
def replace_axi2bram_param(axi2axi_path, axi2axi_cnt, sv_dir, top_name):
    replace(sv_dir+top_name+".sv", "ADDR_WIDTH"      , str(AXI2BRAM_ADDR_WIDTH))
    replace(sv_dir+top_name+".sv", "S_DATA_WIDTH"    , str(AXI2BRAM_S_DATA_WIDTH))
    replace(sv_dir+top_name+".sv", "S_ID_WIDTH"      , str(AXI2BRAM_S_ID_WIDTH))
    replace(sv_dir+top_name+".sv", "S_AWUSER_WIDTH"  , str(AXI2BRAM_S_AWUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "S_WUSER_WIDTH"   , str(AXI2BRAM_S_WUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "S_BUSER_WIDTH"   , str(AXI2BRAM_S_BUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "S_ARUSER_WIDTH"  , str(AXI2BRAM_S_ARUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "S_RUSER_WIDTH"   , str(AXI2BRAM_S_RUSER_WIDTH))   
    
    
#################################################################
# Description    :                                              #
#                                                               #
#                                                               #
# Parameters     :                                              #
#################################################################
def replace_axi2axi_lite_param(axi2axi_path, axi2axi_cnt, sv_dir, top_name):
    replace(sv_dir+top_name+".sv", "ADDR_WIDTH"      , str(AXI2AXI_lite_ADDR_WIDTH))
    replace(sv_dir+top_name+".sv", "S_DATA_WIDTH"    , str(AXI2AXI_lite_S_DATA_WIDTH))
    replace(sv_dir+top_name+".sv", "S_ID_WIDTH"      , str(AXI2AXI_lite_S_ID_WIDTH))
    replace(sv_dir+top_name+".sv", "S_AWUSER_WIDTH"  , str(AXI2AXI_lite_S_AWUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "S_WUSER_WIDTH"   , str(AXI2AXI_lite_S_WUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "S_BUSER_WIDTH"   , str(AXI2AXI_lite_S_BUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "S_ARUSER_WIDTH"  , str(AXI2AXI_lite_S_ARUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "S_RUSER_WIDTH"   , str(AXI2AXI_lite_S_RUSER_WIDTH))
    replace(sv_dir+top_name+".sv", "M_DATA_WIDTH"    , str(AXI2AXI_lite_M_DATA_WIDTH))

#################################################################
# Description    :                                              #
#                                                               #
#                                                               #
# Parameters     :                                              #
#################################################################

def replace_ahb2apb_param(ahb2apb_fabric_path, ahb2apb_cnt, sv_dir, top_name):
    replace(sv_dir+top_name+".sv", "AHB_AW", str(AHB_AW))
    replace(sv_dir+top_name+".sv", "APB_AW", str(APB_AW))
    replace(sv_dir+top_name+".sv", "AHB_DW", str(AHB_DW))
    replace(sv_dir+top_name+".sv", "APB_DW", str(APB_DW))
    replace(sv_dir+top_name+".sv", "FIFO_SIZE", str(FIFO_SIZE))
    replace(sv_dir+top_name+".sv", "AHB_ENDIAN", str(AHB_ENDIAN))
    replace(sv_dir+top_name+".sv", "APB_ENDIAN", str(APB_ENDIAN)) 

#################################################################
# Description    :                                              #
#                                                               #
#                                                               #
# Parameters     :                                              #
#################################################################

def replace_ahb2sram_param(ahb2sram_path, ahb2sram_cnt, sv_dir, top_name):
    replace(sv_dir+top_name+".sv", "DATA_WIDTH", str(DATA_WIDTH))
    replace(sv_dir+top_name+".sv", "MEM_DEPTH", str(MEM_DEPTH)) 

#################################################################
# Description    :                                              #
#                                                               #
#                                                               #
# Parameters     :                                              #
#################################################################
def replace_wire_param(is_wire, axi_fabric_active, ahb_fabric_active, axi2apb_active, axi2axi_active, axi2axi_lite_active, axi2bram_active, ahb2apb_active, ahb2sram_active, sv_dir):
    if is_wire == True and axi_fabric_active:
        replace(sv_dir+"TOP_wire.sv", "DATA_WIDTH", str(DATA_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "ADDR_WIDTH", str(ADDR_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "STRB_WIDTH", str(STRB_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "AW_USER_WIDTH", str(AW_USER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "W_USER_WIDTH", str(W_USER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "B_USER_WIDTH", str(B_USER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "AR_USER_WIDTH", str(AR_USER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "R_USER_WIDTH", str(R_USER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "S_ID_WIDTH",   str(S_ID_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "M_ID_WIDTH",   str(M_ID_WIDTH))
    elif is_wire and ahb_fabric_active:
        replace(sv_dir+"TOP_wire.sv", "AHB_AW", str(AHB_AW))
        replace(sv_dir+"TOP_wire.sv", "AHB_DW", str(AHB_DW))
        replace(sv_dir+"TOP_wire.sv", "NUM_AHB_MASTERS", str(NUM_AHB_MASTERS))
        replace(sv_dir+"TOP_wire.sv", "NUM_AHB_SLAVES", str(NUM_AHB_SLAVES))
        replace(sv_dir+"TOP_wire.sv", "NUM_APB_SLAVES", str(NUM_APB_SLAVES))
        replace(sv_dir+"TOP_wire.sv", "NUM_BEATS_PER_M", str(NUM_BEATS_PER_M))
        replace(sv_dir+"TOP_wire.sv", "NUM_BEATS_PER_M_MT", str(NUM_BEATS_PER_M_MT))
    elif is_wire == True and axi2apb_active:
        replace(sv_dir+"TOP_wire.sv", "DATA_WIDTH", str(AXI_DW))
        replace(sv_dir+"TOP_wire.sv", "ADDR_WIDTH", str(AXI_AW))
        replace(sv_dir+"TOP_wire.sv", "STRB_WIDTH", str(AXI_STRB))
        replace(sv_dir+"TOP_wire.sv", "AW_USER_WIDTH", str(AXI_USER))
        replace(sv_dir+"TOP_wire.sv", "W_USER_WIDTH", str(AXI_USER))
        replace(sv_dir+"TOP_wire.sv", "B_USER_WIDTH", str(AXI_USER))
        replace(sv_dir+"TOP_wire.sv", "AR_USER_WIDTH", str(AXI_USER))
        replace(sv_dir+"TOP_wire.sv", "R_USER_WIDTH", str(AXI_USER))
        replace(sv_dir+"TOP_wire.sv", "S_ID_WIDTH",   str(AXI_IDW))
        replace(sv_dir+"TOP_wire.sv", "M_ID_WIDTH",   str(AXI_IDW))
    elif is_wire == True and axi2axi_active:
        replace(sv_dir+"TOP_wire.sv", "DATA_WIDTH", str(AXI2AXI_S_DATA_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "ADDR_WIDTH", str(AXI2AXI_ADDR_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "STRB_WIDTH", str(int(AXI2AXI_S_DATA_WIDTH/8)))
        replace(sv_dir+"TOP_wire.sv", "AW_USER_WIDTH", str(AXI2AXI_S_AWUSER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "W_USER_WIDTH", str(AXI2AXI_S_WUSER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "B_USER_WIDTH", str(AXI2AXI_S_BUSER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "AR_USER_WIDTH", str(AXI2AXI_S_ARUSER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "R_USER_WIDTH", str(AXI2AXI_S_RUSER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "M_ID_WIDTH",   str(AXI2AXI_S_ID_WIDTH))
    elif is_wire == True and axi2axi_lite_active:
        replace(sv_dir+"TOP_wire.sv", "DATA_WIDTH", str(AXI2AXI_lite_S_DATA_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "ADDR_WIDTH", str(AXI2AXI_lite_ADDR_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "STRB_WIDTH", str(int(AXI2AXI_lite_S_DATA_WIDTH/8)))
        replace(sv_dir+"TOP_wire.sv", "AW_USER_WIDTH", str(AXI2AXI_lite_S_AWUSER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "W_USER_WIDTH", str(AXI2AXI_lite_S_WUSER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "B_USER_WIDTH", str(AXI2AXI_lite_S_BUSER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "AR_USER_WIDTH", str(AXI2AXI_lite_S_ARUSER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "R_USER_WIDTH", str(AXI2AXI_lite_S_RUSER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "M_ID_WIDTH",   str(AXI2AXI_lite_S_ID_WIDTH))
    elif is_wire == True and axi2bram_active:
        replace(sv_dir+"TOP_wire.sv", "S_ID_WIDTH", str(AXI2BRAM_S_ID_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "ADDR_WIDTH", str(AXI2BRAM_ADDR_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "S_DATA_WIDTH", str(AXI2BRAM_S_DATA_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "S_AWUSER_WIDTH", str(AXI2BRAM_S_AWUSER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "S_WUSER_WIDTH", str(AXI2BRAM_S_WUSER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "S_BUSER_WIDTH", str(AXI2BRAM_S_BUSER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "S_ARUSER_WIDTH", str(AXI2BRAM_S_ARUSER_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "S_RUSER_WIDTH",   str(AXI2BRAM_S_RUSER_WIDTH))
    elif is_wire and ahb2apb_active:
        replace(sv_dir+"TOP_wire.sv", "AHB_AW", str(AHB_AW))
        replace(sv_dir+"TOP_wire.sv", "AHB_DW", str(AHB_DW))
        replace(sv_dir+"TOP_wire.sv", "AHB_DW", str(AHB_DW))
        replace(sv_dir+"TOP_wire.sv", "APB_DW", str(APB_DW))
        replace(sv_dir+"TOP_wire.sv", "FIFO_SIZE", str(FIFO_SIZE))
        replace(sv_dir+"TOP_wire.sv", "AHB_ENDIAN", str(AHB_ENDIAN))
        replace(sv_dir+"TOP_wire.sv", "APB_ENDIAN", str(APB_ENDIAN))
    elif is_wire and ahb2sram_active:
        replace(sv_dir+"TOP_wire.sv", "DATA_WIDTH", str(DATA_WIDTH))
        replace(sv_dir+"TOP_wire.sv", "MEM_DEPTH", str(MEM_DEPTH))    

#################################################################
# Description    : Finds all files including folders and        #
#                  subfolders and writes to list                #
# Parameters     :                                              #
#   -des_path    : The path to the file                         #
#                                                               #
# Returns        : List of files                                #
#################################################################
def list_dir(des_path):
    for x in os.listdir(des_path):
        d = os.path.join(des_path, x)
        f_list.append(d)
        if os.path.isdir(d):
            list_dir(d)
    return f_list

#################################################################
# Description    : Environment                                  #
#################################################################
env = Environment(
    loader=FileSystemLoader("templates"),
    autoescape=select_autoescape(
        disabled_extensions=(".txt", ".sv", ".v"),
        default=False
    )
)

#################################################################
# Description    :                                              #
#################################################################
if __name__ == '__main__':

    parser = argparse.ArgumentParser(description="Xlsx parser script",
                                    formatter_class=argparse.ArgumentDefaultsHelpFormatter)
    parser.add_argument("-x", "--xlsx", help="Path to the *.xlsx file")
    parser.add_argument("-t", "--top", help="TOP file name")
    args = parser.parse_args()
    #config = vars(args)

    if args.xlsx == None:
        print("Add the path to the excel file.")
        exit()
    elif not ".xlsx" in args.xlsx:
        print("The argument must have a *.xlsx extension.")
        exit()

    if args.top == None:
        print("Specify the TOP file name (no extension).")
        exit()
    elif "." in args.top:
        print("'.' is not allowed in the name.")
        exit()
    else:
        top_name = args.top

    top_dir = config.RTL_PATH+top_name+"_wrapper/"
    sv_dir = (top_dir+"/sv/")
    top_file = sv_dir+top_name+".sv"
    if os.path.exists(top_dir) == False:
        os.mkdir(top_dir)
    else:
        print("A folder with this TOP name already exists. Delete it or add another name.")
        exit()

    if os.path.exists(sv_dir) == False:
        os.mkdir(sv_dir)
    else:
        print("A folder with this name already exists. Delete it or add another name.")
        exit()

    wb = load_workbook(args.xlsx)
    connections = wb["connections"]

    # If TOP file already exists, delete it and create a new one
    top_file_exists = os.path.exists(top_file)
    if top_file_exists == True:
        os.remove(sv_dir+top_name+".sv")
    add_interface_to_top(sv_dir, top_name, 0, "top_header_template.txt")
    # If the wire file already exists, delete it and create a new
    wire_exists = os.path.exists(sv_dir+"TOP_wire.sv")
    if wire_exists == True:
        os.remove(sv_dir+"TOP_wire.sv")

    # Pick up the NOC number and value from column A
    noc_num = 0
    for x in connections['A']:
        if x.value != None:
            noc_num = noc_num + 1
        else:
            break
    noc_num = noc_num - 1

    NOCs = []
    for x in connections['A']:
        if x.value != "NOC list" :
            NOCs.append(x.value)
    # Generation of TOP file using all NOCs
    mst_con = []
    slv_con = []
    con_cnt = 0
    apb_cnt = 0
    axi_cnt = 0
    ahb_cnt = 0                           
    axi2axi_cnt = 0
    axi2axi_lite_cnt = 0
    axi2bram_cnt = 0
    ahb2apb_cnt = 0
    ahb2sram_cnt = 0                         
    axi_fabric_active = 0
    ahb_fabric_active = 0
    axi2apb_active = 0
    axi2axi_active = 0
    axi2axi_lite_active = 0
    axi2bram_active = 0
    ahb2apb_active = 0
    ahb2sram_active = 0
    for b in range (0, noc_num):
        is_wire = False
        # Add clock and reset for each NOC
        add_clk_rst_to_interface(sv_dir, top_name, NOCs[0])
        # Collect connections for all NOCs
        mst_con, is_wire = get_mst_con(sv_dir, con_cnt, connections, config.mst_offset, top_name, config.first_pos, config.max_con_num, NOCs[0])
        slv_con = get_slv_con(sv_dir, con_cnt, connections, config.slv_offset, top_name, config.first_pos, config.max_con_num, NOCs[0])
        mst_con.reverse()
        slv_con.reverse()
        # Open the current NOC sheet
        sheet = wb[NOCs[0]]
        if "AXI_fabric" in NOCs[0]:
            axi_fabric_active = 1
            axi_fabric_path = sv_dir+NOCs[0]
            # Copies the original AXI files and adapts them
            copy_and_update_axi_file(config.axi_org_path, axi_fabric_path, axi_cnt)
            # Generate the axi_fabric module based on the parameters from Excel
            gen_axi_fabric_by_sheet(sheet, axi_fabric_path,"axi_addr_dec_"+str(axi_cnt), NOCs[0]+"_wrapper",
                                    mst_con, slv_con, config.axi_options_def, config.axi_param_def,
                                    config.first_pos_of_axi_param[axi_cnt], config.param_number[axi_cnt],
                                    config.param_offset[axi_cnt], config.param_value_offset[axi_cnt],
                                    config.addr_first_pos[axi_cnt], config.slv_addr_offset[axi_cnt],
                                    config.start_addr_offset[axi_cnt], config.end_addr_offset[axi_cnt],
                                    config.vis_first_pos[axi_cnt], config.mst_vis_offset[axi_cnt], config.slv_vis_offset[axi_cnt])

            # Replace parameter with value (axi interface)
            replace_axi_fabric_param(axi_fabric_path, axi_cnt, sv_dir, top_name)

            axi_cnt = axi_cnt + 1
        elif "AHB_fabric" in NOCs[0]:
            ahb_fabric_active = 1
            ahb_fabric_path=sv_dir+NOCs[0]
            # Copies the original AHB files and adapts them
            copy_and_update_ahb_file(config.ahb_org_path, ahb_fabric_path, ahb_cnt)
            slaves_range_min, slaves_range_max = read_excel_values(sheet, config.first_pos_of_ahb_param[ahb_cnt], config.ahb_param_number[ahb_cnt], 
                                                                   config.ahb_param_offset[ahb_cnt], config.addr_first_pos[ahb_cnt],
                                                                   config.slv_addr_offset[ahb_cnt], config.start_addr_offset[ahb_cnt], config.end_addr_offset[ahb_cnt])
            # Generate the AHB fabric module based on the parameters from Excel
            gen_ahb_fabric_by_sheet(sheet, ahb_fabric_path, "ahb_addr_dec"+str(ahb_cnt), NOCs[0]+"_wrapper",
                                    mst_con, slv_con, config.ahb_options_def, config.ahb_param_def, config.first_pos_of_ahb_param[ahb_cnt],
                                    config.ahb_param_number[ahb_cnt], config.ahb_param_offset[ahb_cnt], config.ahb_param_value_offset[ahb_cnt], config.addr_first_pos[ahb_cnt], 
                                    config.slv_addr_offset[ahb_cnt],config.start_addr_offset[ahb_cnt], config.end_addr_offset[ahb_cnt],
                                    config.vis_first_pos[ahb_cnt], config.mst_vis_offset[ahb_cnt], config.slv_vis_offset[ahb_cnt], slaves_range_min, slaves_range_max)
            # Replace parameter with value (ahb interface)
            replace_ahb_fabric_param(ahb_fabric_path, ahb_cnt, sv_dir, top_name)

            ahb_cnt += 1
        elif "axi2apb" in NOCs[0]:
            axi2apb_active = 1
            axi2apb_path = sv_dir+NOCs[0]
            # Copies the original AXI files and adapts them
            copy_and_update_apb_file(config.apb_org_path, axi2apb_path, apb_cnt)
            # Generate the axi2apb module based on the parameters from Excel
            gen_axi2apb_by_sheet(sheet, axi2apb_path,"apb_addr_dec_"+str(apb_cnt), NOCs[0]+"_wrapper", 'A',
                                 slv_con, mst_con, config.apb_options_def, config.apb_param_def,
                                 config.first_pos_of_apb_param[apb_cnt], config.apb_param_number[apb_cnt],
                                 config.apb_param_offset[apb_cnt], config.apb_param_value_offset[apb_cnt],
                                 config.addr_first_pos[apb_cnt], config.slv_addr_offset[apb_cnt],
                                 config.start_addr_offset[apb_cnt], config.end_addr_offset[apb_cnt],
                                 config.vis_first_pos[apb_cnt], config.mst_vis_offset[apb_cnt], config.slv_vis_offset[apb_cnt])

            # Replace parameter with value (apb interface)
            replace_axi2apb_param(axi2apb_path, apb_cnt, sv_dir, top_name)

            apb_cnt = apb_cnt + 1
        elif "axi2axi_lite" in NOCs[0]:
            axi2axi_lite_active = 1
            axi2axi_lite_path = sv_dir+NOCs[0]
            os.mkdir(axi2axi_lite_path)
            # Generate the axi2axi_lite module based on the parameters from Excel
            gen_axi2axi_lite_by_sheet(sheet, axi2axi_lite_path, NOCs[0]+"_wrapper", 'A',
                                 slv_con, mst_con, config.axi2axi_lite_options_def, config.axi2axi_lite_param_def,
                                 config.first_pos_of_axi2axi_lite_param[axi2axi_lite_cnt], config.axi2axi_lite_param_number[axi2axi_lite_cnt],
                                 config.axi2axi_lite_param_offset[axi2axi_lite_cnt], config.axi2axi_lite_param_value_offset[axi2axi_lite_cnt])

            
            # Replace parameter with value (axi interface)
            replace_axi2axi_lite_param(axi2axi_lite_path, axi2axi_lite_cnt, sv_dir, top_name)

            axi2axi_lite_cnt = axi2axi_lite_cnt + 1            
        elif "axi2axi" in NOCs[0]:
            axi2axi_active = 1
            axi2axi_path = sv_dir+NOCs[0]
            os.mkdir(axi2axi_path)
            # Generate the axi2axi module based on the parameters from Excel
            gen_axi2axi_by_sheet(sheet, axi2axi_path, NOCs[0]+"_wrapper", 'A',
                                 slv_con, mst_con, config.axi2axi_options_def, config.axi2axi_param_def,
                                 config.first_pos_of_axi2axi_param[axi2axi_cnt], config.axi2axi_param_number[axi2axi_cnt],
                                 config.axi2axi_param_offset[axi2axi_cnt], config.axi2axi_param_value_offset[axi2axi_cnt])

            # Replace parameter with value (axi interface)
            replace_axi2axi_param(axi2axi_path, axi2axi_cnt, sv_dir, top_name)

            axi2axi_cnt = axi2axi_cnt + 1
        elif "axi2bram" in NOCs[0]:
            axi2bram_active = 1
            axi2bram_path = sv_dir+NOCs[0]
            os.mkdir(axi2bram_path)
            # Generate the axi2axi module based on the parameters from Excel
            gen_axi2bram_by_sheet(sheet, axi2bram_path, NOCs[0]+"_wrapper", 'A',
                                 slv_con, mst_con, config.axi2bram_options_def, config.axi2bram_param_def,
                                 config.first_pos_of_axi2bram_param[axi2bram_cnt], config.axi2bram_param_number[axi2bram_cnt],
                                 config.axi2bram_param_offset[axi2bram_cnt], config.axi2bram_param_value_offset[axi2bram_cnt])

            # Replace parameter with value (axi interface)
            replace_axi2bram_param(axi2bram_path, axi2bram_cnt, sv_dir, top_name)

            axi2bram_cnt = axi2bram_cnt + 1 
        elif "ahb2apb" in NOCs[0]:
            ahb2apb_active = 1
            ahb2apb_path = sv_dir+NOCs[0]
            os.mkdir(ahb2apb_path)
            # Generate the ahb2apb module based on the parameters from Excel
            gen_ahb2apb_by_sheet(sheet, ahb2apb_path, NOCs[0]+"_wrapper", 'A',slv_con, mst_con, config.ahb2apb_options_def, config.ahb2apb_param_def,
                                 config.first_pos_of_ahb2apb_param[ahb2apb_cnt], config.ahb2apb_param_number[ahb2apb_cnt],
                                 config.ahb2apb_param_offset[ahb2apb_cnt], config.ahb2apb_param_value_offset[ahb2apb_cnt])

            # Replace parameter with value (ahb2apb interface)
            replace_ahb2apb_param(ahb2apb_path, ahb2apb_cnt, sv_dir, top_name)

            ahb2apb_cnt = ahb2apb_cnt + 1
        elif "ahb2sram" in NOCs[0]:
            ahb2sram_active = 1
            ahb2sram_path = sv_dir+NOCs[0]
            os.mkdir(ahb2sram_path)
            # Generate the ahb2apb module based on the parameters from Excel
            gen_ahb2sram_by_sheet(sheet, ahb2sram_path, NOCs[0]+"_wrapper", 'A',slv_con, mst_con, config.ahb2sram_options_def, config.ahb2sram_param_def,
                                 config.first_pos_of_ahb2sram_param[ahb2sram_cnt], config.ahb2sram_param_number[ahb2sram_cnt],
                                 config.ahb2sram_param_offset[ahb2sram_cnt], config.ahb2sram_param_value_offset[ahb2sram_cnt])

            # Replace parameter with value (ahb2sram interface)
            replace_ahb2sram_param(ahb2sram_path, ahb2sram_cnt, sv_dir, top_name)

            ahb2sram_cnt = ahb2sram_cnt + 1         
        else:
            print("Wrong sheet")

        # Replace parameter with value (wire)
        replace_wire_param(is_wire, axi_fabric_active, ahb_fabric_active,axi2apb_active, axi2axi_active, axi2axi_lite_active, axi2bram_active, ahb2apb_active, ahb2sram_active, sv_dir)
        
        con_cnt = con_cnt + 1
        mst_con.clear()
        slv_con.clear()
        NOCs.remove(NOCs[0])
        axi_fabric_active = 0
        ahb_fabric_active = 0
        axi2apb_active = 0
        axi2axi_active = 0
        axi2axi_lite_active = 0
        axi2bram_active = 0
        ahb2apb_active = 0
        ahb2sram_active = 0

    # Add "end of interface"
    end_interface(sv_dir, top_name)

    # Check if there are wires or just input and output
    wire_exists = os.path.exists(sv_dir+"TOP_wire.sv")

    # Opening temporary files and copying to TOP file
    with open(sv_dir+top_name+".sv", 'a') as c:
        if wire_exists == True:
            with open(sv_dir+'TOP_wire.sv', 'r') as f:
                for line in f:
                    c.write(line)

        for file in os.scandir(sv_dir):
            if file.is_file():
                if "inst_" in file.name:
                    with open(sv_dir+file.name, 'r') as f:
                        for line in f:
                            c.write(line)

        # Adding a endmodule label
        c.write("endmodule")

    # Delete all temporary files
    for x in os.listdir(sv_dir):
        if "_wire" in x or "inst_" in x:
            os.remove(sv_dir+x)

    # Creating a filelist
    with open(top_dir+top_name+"_wrapper_filelist.f", 'w') as f:
        f_list = []
        f_list = list_dir(sv_dir)
        for x in f_list:
            if ".sv" in x:
                f.write(x+"\n")
    replace(top_dir+top_name+"_wrapper_filelist.f", sv_dir, "$RTL_PATH/"+top_name+"_wrapper/sv/")
